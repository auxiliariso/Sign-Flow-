"""
signers/xlsx_signer.py
======================
Firmador para hojas de cálculo Excel (.xlsx).

Capacidades:
  - Detecta placeholders: {{SIGNFLOW}}, {{FIRMA_RESPONSABLE}}, {{MULTIFIRMA}},
    {{RESPONSABLES}} y texto como "Firma:", "Responsable:", "Revisó:", etc.
  - Cada firma agrega una NUEVA FILA en la zona detectada (no sobrescribe).
  - Si no encuentra placeholder, crea o actualiza la hoja "Firmas Digitales".
  - Conserva merges, bordes, estilos y tipografías del documento original.
  - Soporta múltiples responsables en filas separadas.
"""
import io
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

from core.signer_base import SignerBase, SignaturePayload, SignResult
from core.hash_service import HashService

BRAND   = "1F3864"
MID     = "2E5090"
LIGHT   = "D6E4F0"
PALE    = "F0F4FA"
WHITE   = "FFFFFF"
DARK    = "1A1A2E"

SIGNATURE_KEYWORDS = [
    "{{signflow}}", "{{firma_responsable}}", "{{multifirma}}",
    "{{responsables}}", "{{firma}}",
    "firma:", "responsable:", "responsables:", "nombre:", "nombre y firma:",
    "revisó:", "autorizó:", "aprobó:", "elaboró:", "verificó:",
]

SHEET_NAME = "Firmas Digitales"


class XlsxSigner(SignerBase):

    def sign(
        self,
        doc_path: str,
        output_path: str,
        payload: SignaturePayload,
        all_previous: list[SignaturePayload],
    ) -> SignResult:
        wb = load_workbook(doc_path)
        inserted = False

        # 1. Buscar placeholder en todas las hojas
        for ws in wb.worksheets:
            if ws.title == SHEET_NAME:
                continue
            result = self._find_placeholder(ws)
            if result:
                row_idx, col_idx = result
                self._insert_signature_row(ws, row_idx, col_idx, payload)
                inserted = True
                break

        # 2. Sin placeholder → hoja acumulativa
        if not inserted:
            self._update_signature_sheet(wb, payload, all_previous)

        wb.save(output_path)
        doc_hash = HashService().file_hash(output_path)
        return SignResult(
            output_path=output_path,
            firma_hash=payload.firma_hash,
            documento_hash_post=doc_hash,
        )

    def detect_signature_zones(self, doc_path: str) -> list[dict]:
        wb = load_workbook(doc_path, read_only=True)
        zones = []
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        txt = cell.value.strip().lower()
                        for kw in SIGNATURE_KEYWORDS:
                            if kw in txt:
                                zones.append({
                                    "sheet": ws.title,
                                    "row": cell.row,
                                    "col": cell.column,
                                    "text": cell.value.strip(),
                                    "keyword": kw,
                                })
        wb.close()
        return zones

    # ── Detección de placeholder ──────────────────────────────────────────────

    def _find_placeholder(self, ws) -> tuple[int, int] | None:
        """Devuelve (row, col) de la primera celda con keyword, o None."""
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    txt = cell.value.strip().lower()
                    if any(kw in txt for kw in SIGNATURE_KEYWORDS):
                        return cell.row, cell.column
        return None

    # ── Inserción inline en zona detectada ───────────────────────────────────

    def _insert_signature_row(self, ws, row_idx: int, col_idx: int,
                               payload: SignaturePayload):
        """
        Inserta la firma en la celda encontrada y expande la tabla
        agregando subfilas si es necesario.
        """
        cell = ws.cell(row=row_idx, column=col_idx)

        # Si la celda tiene placeholder exacto, la reemplazamos
        txt = (cell.value or "").strip().lower()
        is_placeholder = any(f"{{{{{k.strip('{}')}}}}}" in txt
                             for k in ["{{signflow}}", "{{multifirma}}",
                                       "{{firma_responsable}}", "{{responsables}}"])

        if is_placeholder:
            # Reemplazar el placeholder con los datos de la firma
            cell.value = payload.nombre_completo
            _style_data_cell(cell, bold=True, hex_fg=DARK)

            # Agregar columnas adyacentes con el resto de los datos
            ws.cell(row=row_idx, column=col_idx + 1).value = payload.nombre_puesto
            ws.cell(row=row_idx, column=col_idx + 2).value = payload.fecha
            ws.cell(row=row_idx, column=col_idx + 3).value = payload.validation_id
            for dc in range(1, 4):
                _style_data_cell(ws.cell(row=row_idx, column=col_idx + dc),
                                 hex_fg=DARK)
        else:
            # Insertar nueva fila debajo de la zona detectada
            ws.insert_rows(row_idx + 1)
            for dc, value in enumerate([
                payload.nombre_completo, payload.nombre_puesto,
                payload.fecha, payload.hora, payload.validation_id,
                f"Hash: {payload.firma_hash_short}",
            ]):
                c = ws.cell(row=row_idx + 1, column=col_idx + dc)
                c.value = value
                _style_data_cell(c, bold=(dc == 0), hex_fg=DARK)

    # ── Hoja acumulativa de firmas ────────────────────────────────────────────

    def _update_signature_sheet(
        self,
        wb,
        payload: SignaturePayload,
        all_previous: list[SignaturePayload],
    ):
        # Crear o recuperar hoja
        if SHEET_NAME in wb.sheetnames:
            ws = wb[SHEET_NAME]
        else:
            ws = wb.create_sheet(SHEET_NAME)
            self._build_sheet_header(ws)

        # Encontrar siguiente fila vacía (después del encabezado)
        next_row = _next_empty_row(ws, start=3)

        # Zebra row
        is_even = (next_row % 2 == 0)
        bg = LIGHT if is_even else PALE

        data = [
            str(payload.id_firma),
            payload.validation_id,
            payload.nombre_completo,
            payload.nombre_puesto,
            payload.fecha,
            payload.hora,
            payload.firma_hash_short,
        ]
        thin = Side(style="thin", color=BRAND)
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col, value in enumerate(data, start=1):
            c = ws.cell(row=next_row, column=col, value=value)
            c.font = Font(name="Calibri", size=9,
                          bold=(col == 3), color=DARK)
            c.fill = PatternFill("solid", fgColor=bg)
            c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=(col == 7))
            ws.row_dimensions[next_row].height = 18

        # QR como imagen en la columna 8
        if payload.qr_image_bytes:
            try:
                img = XLImage(io.BytesIO(payload.qr_image_bytes))
                img.width  = 48
                img.height = 48
                col_letter = get_column_letter(8)
                ws.add_image(img, f"{col_letter}{next_row}")
                ws.row_dimensions[next_row].height = 40
            except Exception:
                pass

    def _build_sheet_header(self, ws):
        """Construye la cabecera de la hoja de firmas."""
        # Fila 1: título
        ws.merge_cells("A1:H1")
        title = ws["A1"]
        title.value = "✦ FIRMAS DIGITALES — SignFlow"
        title.font  = Font(name="Calibri", bold=True, size=13,
                           color=WHITE)
        title.fill  = PatternFill("solid", fgColor=BRAND)
        title.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        # Fila 2: columnas
        headers = ["#", "ID Validación", "Firmado por", "Cargo",
                   "Fecha", "Hora", "Hash", "QR"]
        widths  = [5, 18, 26, 22, 12, 10, 20, 8]
        thin = Side(style="thin", color=BRAND)
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col, (h, w) in enumerate(zip(headers, widths), start=1):
            c = ws.cell(row=2, column=col, value=h)
            c.font      = Font(name="Calibri", bold=True, size=10, color=WHITE)
            c.fill      = PatternFill("solid", fgColor=MID)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = border
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[2].height = 20


# ── helpers ───────────────────────────────────────────────────────────────────

def _style_data_cell(cell, bold=False, hex_fg=DARK, bg=PALE):
    thin = Side(style="thin", color=BRAND)
    cell.font      = Font(name="Calibri", bold=bold, size=9, color=hex_fg)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell.alignment = Alignment(vertical="center")


def _next_empty_row(ws, start: int = 3) -> int:
    for row in range(start, ws.max_row + 2):
        if ws.cell(row=row, column=1).value is None:
            return row
    return ws.max_row + 1
