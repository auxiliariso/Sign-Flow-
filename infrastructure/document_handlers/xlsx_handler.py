"""
Handler - Firma en hojas de cálculo Excel (.xlsx)
Agrega una hoja "Firma Digital" con el registro de firma.
"""
import shutil
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from core.interfaces.repositories import IDocumentSigner

BRAND_HEX = "1F3864"
LIGHT_BLUE = "D6E4F0"


class XlsxSigner(IDocumentSigner):
    def sign(
        self,
        doc_path: str,
        output_path: str,
        nombre_completo: str,
        nombre_puesto: str,
        firma_hash: str,
        fecha: str,
        hora: str,
    ) -> str:
        shutil.copy2(doc_path, output_path)
        wb = openpyxl.load_workbook(output_path)

        # Eliminar hoja previa de firma si existe
        if "Firma Digital" in wb.sheetnames:
            del wb["Firma Digital"]

        ws = wb.create_sheet("Firma Digital")
        _build_sheet(ws, nombre_completo, nombre_puesto, firma_hash, fecha, hora)

        wb.save(output_path)
        return output_path


def _build_sheet(ws, nombre, puesto, hash_val, fecha, hora):
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill("solid", fgColor=BRAND_HEX)
    label_font  = Font(name="Calibri", bold=True, color=BRAND_HEX, size=10)
    value_font  = Font(name="Calibri", size=10)
    alt_fill    = PatternFill("solid", fgColor=LIGHT_BLUE)
    thin        = Side(style="thin", color=BRAND_HEX)
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    center      = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 64

    # Título
    ws.merge_cells("A1:B1")
    title_cell = ws["A1"]
    title_cell.value = "✦ Firma Digital — SignFlow"
    title_cell.font = header_font
    title_cell.fill = header_fill
    title_cell.alignment = center
    title_cell.border = border
    ws.row_dimensions[1].height = 26

    rows = [
        ("Firmado por", nombre),
        ("Cargo / Puesto", puesto),
        ("Fecha (UTC)", fecha),
        ("Hora (UTC)", hora),
        ("Hash de verificación", hash_val),
    ]

    for i, (label, value) in enumerate(rows, start=2):
        lc = ws.cell(row=i, column=1, value=label)
        vc = ws.cell(row=i, column=2, value=value)
        lc.font = label_font
        vc.font = value_font
        for cell in (lc, vc):
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if i % 2 == 0:
                cell.fill = alt_fill
        ws.row_dimensions[i].height = 18
